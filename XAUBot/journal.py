"""
================================================================================
GOLDBOT - Trading Journal
Otomatis catat setiap trade ke Excel jurnal harian
File: GoldBot/journal.py
================================================================================
"""

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ============================================================================
# KONSTANTA STYLE
# ============================================================================

# Header
HDR_BG   = "1F3864"   # navy dark
HDR_FG   = "FFFFFF"   # putih
# Summary row
SUM_BG   = "D6E4F0"   # biru muda
# Row warna
WIN_BG   = "E2EFDA"   # hijau muda
LOSS_BG  = "FCE4D6"   # merah muda
ALT_BG   = "F2F2F2"   # abu ringan

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADERS = [
    "#", "Tanggal", "Waktu Open", "Waktu Close",
    "Durasi (menit)", "Simbol", "Arah", "Lot",
    "Harga Open", "Harga Close", "SL", "TP",
    "Profit ($)", "Komisi ($)", "Net P&L ($)",
    "Exit Reason", "Kumulatif ($)", "Catatan"
]

COL_WIDTHS = [5, 13, 13, 13, 16, 10, 7, 6,
              12, 12, 10, 10,
              12, 12, 12,
              13, 14, 20]


def _thin_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _apply_header(ws):
    for col_idx, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = Font(bold=True, color=HDR_FG, name="Arial", size=10)
        cell.fill      = PatternFill("solid", fgColor=HDR_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def _col(name: str) -> int:
    return HEADERS.index(name) + 1


# ============================================================================
# SUMMARY SHEET
# ============================================================================

def _update_summary(wb, trades_ws_name: str):
    """Buat/update sheet Summary dengan statistik harian."""
    if "Summary" in wb.sheetnames:
        del wb["Summary"]

    ws = wb.create_sheet("Summary", 0)

    # Judul
    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value     = "📊 GoldBot — Ringkasan Performa"
    title.font      = Font(bold=True, size=14, name="Arial", color=HDR_FG)
    title.fill      = PatternFill("solid", fgColor=HDR_BG)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    labels = [
        ("", ""),
        ("Total Trade",         f"=COUNTA('{trades_ws_name}'!A2:A10000)-COUNTBLANK('{trades_ws_name}'!A2:A10000)"),
        ("Trade WIN",           f"=COUNTIF('{trades_ws_name}'!O2:O10000,\">0\")"),
        ("Trade LOSS",          f"=COUNTIF('{trades_ws_name}'!O2:O10000,\"<0\")"),
        ("Win Rate (%)",        f"=IFERROR(C4/C3*100,0)"),
        ("",                    ""),
        ("Total Net P&L ($)",   f"=SUMIF('{trades_ws_name}'!A2:A10000,\"<>\",'{trades_ws_name}'!O2:O10000)"),
        ("Best Trade ($)",      f"=IFERROR(MAX('{trades_ws_name}'!O2:O10000),0)"),
        ("Worst Trade ($)",     f"=IFERROR(MIN('{trades_ws_name}'!O2:O10000),0)"),
        ("Avg Win ($)",         f"=IFERROR(AVERAGEIF('{trades_ws_name}'!O2:O10000,\">0\"),0)"),
        ("Avg Loss ($)",        f"=IFERROR(AVERAGEIF('{trades_ws_name}'!O2:O10000,\"<0\"),0)"),
        ("Profit Factor",       f"=IFERROR(SUMIF('{trades_ws_name}'!O2:O10000,\">0\")/ABS(SUMIF('{trades_ws_name}'!O2:O10000,\"<0\")),0)"),
        ("",                    ""),
        ("Updated",             f"=NOW()"),
    ]

    for r, (lbl, val) in enumerate(labels, start=2):
        if not lbl:
            continue
        lc = ws.cell(row=r, column=2, value=lbl)
        vc = ws.cell(row=r, column=3, value=val)
        lc.font = Font(name="Arial", size=10, bold=True)
        vc.font = Font(name="Arial", size=10)
        lc.border = _thin_border()
        vc.border = _thin_border()
        lc.fill  = PatternFill("solid", fgColor=SUM_BG)
        vc.alignment = Alignment(horizontal="right")

        # Format khusus
        if "%" in lbl:
            vc.number_format = "0.0%"
        elif "$" in lbl:
            vc.number_format = '#,##0.00'
        elif "Factor" in lbl:
            vc.number_format = "0.00"
        elif "Updated" in lbl:
            vc.number_format = "DD/MM/YYYY HH:MM"

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 3


# ============================================================================
# TRADING JOURNAL CLASS
# ============================================================================

class TradingJournal:
    """
    Jurnal trading otomatis ke Excel.

    Cara pakai di bot.py:
        from journal import TradingJournal
        self.journal = TradingJournal(config.journal_file)

    Saat trade close:
        self.journal.log_trade(entry)
    """

    def __init__(self, filepath: str):
        self.filepath = filepath
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        self._ensure_file()

    # ------------------------------------------------------------------
    # FILE MANAGEMENT
    # ------------------------------------------------------------------

    def _sheet_name(self) -> str:
        """Nama sheet = bulan berjalan, e.g. 'Apr 2026'."""
        return datetime.now().strftime("%b %Y")

    def _ensure_file(self):
        """Buat file + sheet bulan ini kalau belum ada."""
        if os.path.exists(self.filepath):
            wb = load_workbook(self.filepath)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        sheet_name = self._sheet_name()
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            _apply_header(ws)

        _update_summary(wb, sheet_name)
        wb.save(self.filepath)

    # ------------------------------------------------------------------
    # LOG TRADE
    # ------------------------------------------------------------------

    def log_trade(self,
                  ticket:       int,
                  symbol:       str,
                  direction:    str,
                  lot:          float,
                  price_open:   float,
                  price_close:  float,
                  sl:           float,
                  tp:           float,
                  profit:       float,
                  exit_reason:  str,
                  open_time:    datetime = None,
                  close_time:   datetime = None,
                  commission:   float    = 0.05,
                  note:         str      = ""):
        """
        Catat satu trade ke jurnal Excel.

        Parameters:
            ticket       : nomor tiket MT5
            symbol       : e.g. "XAUUSD.m"
            direction    : "BUY" atau "SELL"
            lot          : volume lot
            price_open   : harga buka posisi
            price_close  : harga tutup posisi
            sl           : stop loss
            tp           : take profit
            profit       : profit gross (dari MT5)
            exit_reason  : "SL" / "TP" / "TRAIL" / "MANUAL" / "SWAP"
            open_time    : datetime buka (default: now)
            close_time   : datetime tutup (default: now)
            commission   : komisi per 0.01 lot = $0.05
            note         : catatan bebas
        """
        now = datetime.now()
        if open_time  is None: open_time  = now
        if close_time is None: close_time = now

        duration_min = max(0, round((close_time - open_time).total_seconds() / 60))
        net_pnl      = profit - abs(commission)

        wb         = load_workbook(self.filepath)
        sheet_name = self._sheet_name()

        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            _apply_header(ws)

        ws       = wb[sheet_name]
        next_row = ws.max_row + 1

        # Nomor urut (skip header)
        trade_num = next_row - 1

        # Kolom kumulatif: jumlah net P&L semua baris sebelumnya + baris ini
        prev_cum_col = get_column_letter(_col("Kumulatif ($)"))
        cum_formula  = (
            f"=SUMIF($A$2:A{next_row-1},\"<>\",$O$2:O{next_row-1})+O{next_row}"
            if next_row > 2
            else f"=O{next_row}"
        )

        row_data = {
            "#":              trade_num,
            "Tanggal":        open_time.strftime("%d/%m/%Y"),
            "Waktu Open":     open_time.strftime("%H:%M:%S"),
            "Waktu Close":    close_time.strftime("%H:%M:%S"),
            "Durasi (menit)": duration_min,
            "Simbol":         symbol,
            "Arah":           direction,
            "Lot":            lot,
            "Harga Open":     price_open,
            "Harga Close":    price_close,
            "SL":             sl,
            "TP":             tp if tp not in (0.01, 99999.0) else None,
            "Profit ($)":     profit,
            "Komisi ($)":     -abs(commission),
            "Net P&L ($)":    net_pnl,
            "Exit Reason":    exit_reason,
            "Kumulatif ($)":  cum_formula,
            "Catatan":        f"#{ticket} {note}".strip(),
        }

        # Tentukan warna baris
        bg = WIN_BG if net_pnl >= 0 else LOSS_BG

        for col_idx, header in enumerate(HEADERS, start=1):
            cell       = ws.cell(row=next_row, column=col_idx, value=row_data[header])
            cell.font  = Font(name="Arial", size=10)
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill  = PatternFill("solid", fgColor=bg)

            # Number format
            if header in ("Profit ($)", "Komisi ($)", "Net P&L ($)", "Kumulatif ($)"):
                cell.number_format = '#,##0.00'
            elif header in ("Harga Open", "Harga Close", "SL", "TP"):
                cell.number_format = '#,##0.00'
            elif header == "Lot":
                cell.number_format = '0.00'

        # Arah cell warna teks
        arah_cell = ws.cell(row=next_row, column=_col("Arah"))
        arah_cell.font = Font(
            name="Arial", size=10, bold=True,
            color="0070C0" if direction == "BUY" else "C00000"
        )

        # Update summary
        _update_summary(wb, sheet_name)
        wb.save(self.filepath)

    # ------------------------------------------------------------------
    # DAILY SUMMARY ROW
    # ------------------------------------------------------------------

    def add_daily_separator(self, daily_pnl: float, trade_count: int):
        """Tambah baris separator ringkasan akhir hari."""
        wb         = load_workbook(self.filepath)
        sheet_name = self._sheet_name()
        if sheet_name not in wb.sheetnames:
            wb.save(self.filepath)
            return

        ws       = wb[sheet_name]
        next_row = ws.max_row + 1
        today    = datetime.now().strftime("%d/%m/%Y")

        label = f"── Ringkasan {today} | {trade_count} trades | Net: ${daily_pnl:+.2f} ──"
        ws.merge_cells(f"A{next_row}:R{next_row}")
        cell = ws.cell(row=next_row, column=1, value=label)
        cell.font      = Font(name="Arial", size=10, bold=True, italic=True, color="595959")
        cell.fill      = PatternFill("solid", fgColor=SUM_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()
        ws.row_dimensions[next_row].height = 18

        _update_summary(wb, sheet_name)
        wb.save(self.filepath)
